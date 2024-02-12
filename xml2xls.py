# This program lists all files in current directory and searches for .xml files
# It then reads the xml files one by one and writes data to the excel table
# Data to read:
# - SSERIENNR1 -> Part No.
# - SSERIENNR2 -> Part No. 2
# - STTIMEAUTOMATIKSTART1 -> Začátek otevíraní formy
# - STTIMEAUTOMATIKEND2 -> Konec zavření formy a celého procesu
# - STTIMEKLEBENEND -> Konec lepení
# - STTIMEKLEBELEHREABKUPPELN -> Odpojení formy od Hartingu
# - STTIMEKLEBENDOPAGEND -> Konec lepení lepidlem 2K
#
# From every variable, read date and time, stored in:
# WYEAR, WMONTH, WDAY, WHOUR, WMINUTE, WSECOND

input_dir = "xml"
output_dir = "xls"
keys = ['SSERIENNR1', 'SSERIENNR2', 'STTIMEAUTOMATIKSTART1', 'STTIMEAUTOMATIKEND2', 'STTIMEKLEBENEND', 'STTIMEKLEBELEHREABKUPPELN', 'STTIMEKLEBENDOPAGEND']
names = ['Part No.', 'Part No. 2', 'Začátek otevíraní formy', 'Konec zavření formy a celého procesu', 'Konec lepení', 'Odpojení formy od Hartingu', 'Konec lepení lepidlem 2K']
offset = 2
# Import libraries
import os
import xml.etree.ElementTree as ET
import openpyxl
from datetime import datetime

# List xml files
def list_files():
	files = []
	for file in os.listdir(input_dir):
		if file.endswith(".xml"):
			files.append(file)
			print("Found ", len(files), "files", end='\r')
	#sort files by date, filename format: XXXXXXX_XXXXXXX_YYYYMMDD_HHMMSS.xml
	print("\nSorting...")
	files.sort(key=lambda x: datetime.strptime(x.split('_')[2] + '_' + x.split('_')[3][0:6], '%Y%m%d_%H%M%S'))
	return files
# Read xml file to object
def read_xml(file):
	tree = ET.parse(input_dir + "/" + file)
	root = tree.getroot()
	item = {}
	item['file'] = file
	# Read items in file
	for items in root:
		#print(items.tag, items.attrib)
		# Read elements in items
		for elements in items:
			#print(elements.tag, elements.attrib)
			if elements.tag in keys:
				# Read date and time
				item[elements.tag] = elements.text
				date = {}
				for properties in elements:
					date[properties.tag] = properties.text
				if date:
					item[elements.tag] = date2timestamp(date)
	
	return item

def date2timestamp(date):
	# Convert date to timestamp
	try:
		date = datetime(int(date['WYEAR']), int(date['WMONTH']), int(date['WDAY']), int(date['WHOUR']), int(date['WMINUTE']), int(date['WSECOND']))#.strftime('%H:%M:%S %d-%m-%Y')
	except:
		date = "00:00:00"
	return date

def excel_date(date1):
	# From: https://stackoverflow.com/questions/9574793/how-to-convert-a-python-datetime-datetime-to-excel-serial-date-number
    temp = datetime(1899, 12, 30)    # Note, not 31st Dec but 30th!
    delta = date1 - temp
    return float(delta.days) + (float(delta.seconds) / 86400)

# Main function
def main():
	print("Starting xml2xls converter...")
	# Check for directories
	if not os.path.exists(input_dir):
		os.makedirs(input_dir)
	if not os.path.exists(output_dir):
		os.makedirs(output_dir)
	# Check if template files exists
	if not os.path.exists('templates/reclamacion_template.xlsx'):
		print("Error: File templates/reclamacion_template.xlsx not found")
		return
	if not os.path.exists('templates/results_template.xlsx'):
		print("Error: File templates/results_template.xlsx not found")
		return
	# List files
	files = list_files()
	# Create workbook for results
	results = openpyxl.load_workbook('templates/results_template.xlsx')
	result_sheet = results.active
	iterations = 0
	last_part_end = 0

	for file in files:
		#print("Processing file %d of %d" % (iterations+1, len(files)), end='\r')
		# Read template excel file
		wb = openpyxl.load_workbook('templates/reclamacion_template.xlsx')
		root = read_xml(file)
		# Select active sheet
		sheet = wb.active
		# Write data to excel
		sheet.cell(row=2, column=2).value = names[0] + " " + root[keys[0]] # Part No.

		for i in range(len(keys)-offset):
			sheet.cell(row=3+i, column=2).value = keys[offset+i]
			sheet.cell(row=3+i, column=3).value = names[offset+i]
			try:
				sheet.cell(row=3+i, column=4).value = excel_date(root[keys[offset+i]])
				sheet.cell(row=3+i, column=4).number_format = 'HH:MM:SS'
			except:
				sheet.cell(row=3+i, column=4).value = root[keys[offset+i]]

		# Save excel file
		try:
			wb.save(output_dir + "/" + file[:-4] + ".xlsx")
			print("File " + file + " processed")
		except:
			print("Error while saving file " + file)

		result_sheet.cell(row=2+iterations, column=1).value = excel_date(root[keys[offset]])
		result_sheet.cell(row=2+iterations, column=1).number_format = 'dd.mm.yyyy'
		result_sheet.cell(row=2+iterations, column=2).value = int(root[keys[0]])
		result_sheet.cell(row=2+iterations, column=3).value = '=HYPERLINK("{}", "{}")'.format(output_dir + "/" + file[:-4] + ".xlsx", file[:-4] + ".xlsx")
		result_sheet.cell(row=2+iterations, column=4).value = excel_date(root[keys[offset+1]]) - excel_date(root[keys[offset+4]])
		result_sheet.cell(row=2+iterations, column=4).number_format = 'HH:MM:SS'
		if iterations == 0:
			last_part_end = excel_date(root[keys[offset+1]])
		else:
			result_sheet.cell(row=2+iterations, column=5).value = excel_date(root[keys[offset]]) - last_part_end
			result_sheet.cell(row=2+iterations, column=5).number_format = 'HH:MM:SS'
			last_part_end = excel_date(root[keys[offset+1]])

		note = ""
		if result_sheet.cell(row=2+iterations, column=4).value < excel_date(datetime(1899, 12, 30, 0, 1, 30)):
			result_sheet.cell(row=2+iterations, column=6).value = "NOK"
			note = "Krátký čas lepení" if note == "" else note + ", Krátký čas lepení"

		elif result_sheet.cell(row=2+iterations, column=4).value > excel_date(datetime(1899, 12, 30, 0, 2, 0)):
			result_sheet.cell(row=2+iterations, column=6).value = "NOK"
			note = "Dlouhý čas lepení" if note == "" else note + ", Dlouhý čas lepení"

		elif iterations > 0 and result_sheet.cell(row=2+iterations, column=5).value > excel_date(datetime(1899, 12, 30, 1, 0, 00)):
			result_sheet.cell(row=2+iterations, column=6).value = "NOK"
			note = "Dlouhá doba mezi díly" if note == "" else note + ", Dlouhá doba mezi díly"
		
		else:
			result_sheet.cell(row=2+iterations, column=6).value = "OK"

		result_sheet.cell(row=2+iterations, column=7).value = note
		iterations += 1
	# Save results
	results.save("results.xlsx")
	print("Processing finished.")



# Run main function
if __name__ == "__main__":
	main()
	input("Press Enter to close the window...")